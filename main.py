import openpyxl


def print_it(job, eff_date, pol, row, last_row):
    entry_screen_flag = "true" if row == 1 else "false"
    exit_screen_flag = "true" if last_row == 1 else "false"
    str_screen_num = (row - 1) * 3

    tab_key = "" if len(job) == 10 else "[tab]"

    ha_script = f"""
    <screen name="Screen{str_screen_num + 1}" entryscreen="{entry_screen_flag}" exitscreen="false" transient="false">
        <description>
            <oia status="NOTINHIBITED" optional="false" invertmatch="false"/>
            <numfields number="101" optional="true" invertmatch="false"/>
            <numinputfields number="4" optional="true" invertmatch="false"/>
        </description>

        <actions>
            <input value="{job}{tab_key}{eff_date}[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true"
                   encrypted="false"/>
        </actions>

        <nextscreens timeout="0">
            <nextscreen name="Screen{str_screen_num + 2}"/>
        </nextscreens>
    </screen>


    <screen name="Screen{str_screen_num + 2}" entryscreen="false" exitscreen="false" transient="false">
        <description>
            <oia status="NOTINHIBITED" optional="false" invertmatch="false"/>
            <numfields number="36" optional="true" invertmatch="false"/>
            <numinputfields number="2" optional="true" invertmatch="false"/>
        </description>

        <actions>
            <input value="{pol}{pol}[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true"
                   encrypted="false"/>
        </actions>

        <nextscreens timeout="0">
            <nextscreen name="Screen{str_screen_num + 3}"/>
        </nextscreens>
    </screen>


    <screen name="Screen{str_screen_num + 3}" entryscreen="false" exitscreen="{exit_screen_flag}" transient="false">
        <description>
            <oia status="NOTINHIBITED" optional="false" invertmatch="false"/>
            <numfields number="36" optional="true" invertmatch="false"/>
             <numinputfields number="1" optional="true" invertmatch="false"/>
        </description>

        <actions>
            <input value="[tab]y[enter]" row="0" col="0" movecursor="true" xlatehostkeys="true" encrypted="false"/>
        </actions>

        <nextscreens timeout="0">
            <nextscreen name="Screen{str_screen_num + 4}"/>
        </nextscreens>
    </screen>
    """
    print(ha_script)


def gen_acs_macro():
    # global job, eff_date, pol
    wb1 = openpyxl.load_workbook("jobs.xlsx")

    wb = wb1.active

    # for row in wb.iter_rows():
    #     for cell in row:
    #         print(cell.value, end=" ")
    #     print("\n")

    # for row in range(wb.max_row):
    #     for col in wb.iter_cols(1, wb.max_column):
    #         print(col[row].value, end=' ')
    #     print()

    # for row in range(1, wb.max_row + 1):
    #     for col in range(1, wb.max_column + 1):
    #         cell = wb.cell(row, col)
    #         if col == 1:
    #             job = cell
    #             print(cell.value, end=' ')
    #         elif col == 2:
    #             eff_date = str(cell.value)[5:7] + '/' + str(cell.value)[8:10] + '/' + str(cell.value)[0:4]
    #             print(eff_date, end=' ')
    #         else:
    #             pol = cell
    #             print(cell.value, end=' ')
    #     print()

    # ha script header
    print(
        '<HAScript name="macro1" description="" timeout="60000" pausetime="300" promptall="true" blockinput="true" author="vpc" creationdate="Nov 15, 2023 8:21:30 AM" supressclearevents="false" usevars="false" ignorepauseforenhancedtn="true" delayifnotenhancedtn="0" ignorepausetimeforenhancedtn="true" continueontimeout="false">')
    for row in range(1, wb.max_row + 1):
        for col in range(1, wb.max_column + 1):
            cell = wb.cell(row, col)
            if col == 1:
                job = cell.value
            elif col == 2:
                eff_date = str(cell.value)[5:7] + '/' + str(cell.value)[8:10] + '/' + str(cell.value)[0:4]
            elif col == 3:
                pol = cell.value
        print_it(job, eff_date, pol, row, row == wb.max_row)

    # ha script footer
    print('</HAScript>')


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    gen_acs_macro()